#!/usr/bin/env python3
"""财务 Agent — 管报底座 + 绩效试算 Demo (D4)"""
from flask import Flask, request, jsonify, render_template_string, redirect
import os
import json
import sqlite3
import io
from dotenv import load_dotenv
from cherry_client import chat, chat_json, embed, test_connection
from feishu_client import list_chats as feishu_list_chats, send_post as feishu_send_post, send_text as feishu_send_text, download_message_file as feishu_download_file
from monitor import init_monitor

load_dotenv()

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "finance-agent-secret-dev")
from flask import session as flask_session
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB

# === 配置 ===
CHERRYIN_API_KEY = os.environ.get("CHERRYIN_API_KEY", "")
CHERRYIN_BASE_URL = os.environ.get("CHERRYIN_BASE_URL", "https://express-ent-admin.cherryin.ai/v1")
LLM_MODEL = os.environ.get("LLM_MODEL", "agent/deepseek-v4-pro")

DB_PATH = os.path.join(os.environ.get("DATA_DIR", os.path.dirname(__file__)), "finance.db")
ALERT_CHAT_ID = os.environ.get("FEISHU_ALERT_CHAT_ID", "")  # 留空=跳过飞书推送

# === 数据库初始化 ===
def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS transactions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        source TEXT,
        amount REAL,
        summary TEXT,
        level1 TEXT,
        level2 TEXT,
        confidence REAL,
        reason TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    # === 权限分层:users 表 ===
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'employee',
        pin TEXT,
        feishu_open_id TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    # 预置账号(首次建表)
    c.execute("SELECT COUNT(*) FROM users")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO users (name, role, pin) VALUES (?, ?, ?)", ("财务管理员", "financial", "1234"))
        c.execute("INSERT INTO users (name, role, pin) VALUES (?, ?, ?)", ("张三", "employee", "1111"))
        c.execute("INSERT INTO users (name, role, pin) VALUES (?, ?, ?)", ("李四", "employee", "2222"))

    # === 发票流改造:给 transactions 表加字段(兼容旧库) ===
    try:
        c.execute("ALTER TABLE transactions ADD COLUMN vendor TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE transactions ADD COLUMN invoice_no TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE transactions ADD COLUMN invoice_text TEXT")
    except Exception:
        pass
    try:
        c.execute("ALTER TABLE transactions ADD COLUMN user_id INTEGER")
    except Exception:
        pass

    # D5: 绩效规则表
    c.execute("""CREATE TABLE IF NOT EXISTS performance_rules (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        department TEXT NOT NULL,
        position TEXT NOT NULL,
        coefficient REAL DEFAULT 1.0,
        target_amount REAL DEFAULT 0,
        bonus_base REAL DEFAULT 0,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )""")
    # 插入默认规则(首次建表时)
    c.execute("SELECT COUNT(*) FROM performance_rules")
    if c.fetchone()[0] == 0:
        default_rules = [
            ("研发部", "工程师", 1.2, 500000, 8000),
            ("研发部", "主管", 1.5, 800000, 15000),
            ("销售部", "销售", 1.0, 300000, 6000),
            ("销售部", "经理", 1.8, 1000000, 20000),
            ("管理部", "行政", 0.8, 200000, 5000),
            ("管理部", "总监", 2.0, 500000, 25000),
            ("交付部", "交付工程师", 1.1, 400000, 7000),
            ("交付部", "主管", 1.4, 700000, 12000),
        ]
        c.executemany("INSERT INTO performance_rules (department, position, coefficient, target_amount, bonus_base) VALUES (?, ?, ?, ?, ?)", default_rules)
    conn.commit()
    conn.close()

init_db()

# === 监控初始化 ===
init_monitor(app, service_name="finance-agent", db_path=DB_PATH, llm_test_fn=test_connection,
             alert_feishu_fn=feishu_send_post, alert_chat_id=ALERT_CHAT_ID)

def _current_user():
    """从 session 取当前用户,未登录返回 None"""
    uid = flask_session.get("user_id")
    if not uid:
        return None
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT id, name, role FROM users WHERE id=?", (uid,))
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return None
    return {"id": row[0], "name": row[1], "role": row[2]}

def _require_user():
    """要求登录,返回 user 或 (None, error_response)"""
    u = _current_user()
    if not u:
        return None, jsonify({"ok": False, "error": "未登录"}), 401
    return u, None


def _get_or_create_user_by_open_id(open_id, name=None):
    """通过飞书 open_id 查或建用户。返回 user dict 或 None。"""
    if not open_id:
        return None
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT id, name, role FROM users WHERE feishu_open_id=?", (open_id,))
        row = c.fetchone()
        if row:
            return {"id": row[0], "name": row[1], "role": row[2]}
        # 自动建档为 employee
        c.execute("INSERT INTO users (name, role, feishu_open_id) VALUES (?, ?, ?)",
                  (name or f"飞书用户_{open_id[-6:]}", "employee", open_id))
        conn.commit()
        uid = c.lastrowid
        return {"id": uid, "name": name or f"飞书用户_{open_id[-6:]}", "role": "employee"}
    finally:
        conn.close()


# === 口径规则 ===

# === OCR 配置(发票扫描件识别) ===
_OCR_MAX_PAGES = 50
_OCR_DPI = 200
_OCR_MIN_CHARS = 20

ACCOUNTING_RULES = """
一、研发费:
  - 研发人员薪酬(直接归属研发项目)
  - 软件开发外包费
  - 开发工具/软件许可
  - 测试费、云服务费(研发用)
  - 研发相关差旅

二、销售费:
  - 销售人员薪酬+提成
  - 拜访客户差旅
  - 客户招待费
  - 市场推广/广告
  - 销售佣金

三、管理费:
  - 行政人员薪酬
  - 办公租金
  - 办公用品
  - 非销售差旅(内部会议、培训)
  - 团队建设
  - 法务/财务/HR 职能费用

四、营业成本:
  - 产品采购成本
  - 外包服务交付成本
  - 交付人员薪酬
"""

# === 归一化 Prompt ===
NORMALIZE_PROMPT = """你是财务归类助手。根据以下流水信息,归一化到标准科目。

## 口径规则
{rules}

## 流水信息
- 金额: {amount} 元
- 摘要: {summary}
- 来源: {source}

## 输出要求
返回 JSON(不要其他文字):
{{
  "level1": "研发费|销售费|管理费|营业成本",
  "level2": "二级科目(如:差旅费/薪酬/软件/招待费等)",
  "confidence": 0.0到1.0,
  "reason": "30字以内判断依据"
}}
"""

# === 发票解析+归一化 Prompt(一步到位) ===
INVOICE_PARSE_PROMPT = """你是财务发票解析助手。从 OCR 提取的发票文本中,提取结构化字段并归一化科目。

## 口径规则
{rules}

## 发票 OCR 文本
{invoice_text}

## 输出要求
返回 JSON(不要其他文字):
{{
  "invoice_no": "发票号码(没找到则空字符串)",
  "invoice_date": "开票日期 YYYY-MM-DD(没找到则空)",
  "vendor": "销售方名称(没找到则空)",
  "amount": 0.0,
  "items": ["商品/服务明细1", "商品/服务明细2"],
  "level1": "研发费|销售费|管理费|营业成本",
  "level2": "二级科目(如:差旅费/薪酬/软件/招待费/办公用品等)",
  "confidence": 0.0到1.0,
  "reason": "30字以内判断依据"
}}

注意:
- amount 是价税合计金额(发票总金额),纯数字不要带¥或元
- 如果有多行商品,items 列出每行
- 如果 OCR 文本残缺无法判断,confidence 给低分(0.3 以下),level1 给最可能猜测
- 如果完全不是发票(如普通文本),level1 给"未归类",confidence 给 0
"""


# === D4: AI 简评 Prompt ===
REPORT_COMMENTARY_PROMPT = """你是财务分析师。根据以下管报汇总数据,写出 3-5 条简短点评。

要求:
- 指出占比最高的科目
- 发现异常波动或值得关注的点
- 给出 1-2 条优化建议
- 每条不超过 50 字
- 用中文,口语化,像同事汇报

管报数据:
{report_data}
"""

# === D4: 管报页面 ===
REPORT_HTML = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>管理报表 · 财务 Agent</title>
<style>
:root{
  --background: oklch(0.145 0.004 270);
  --foreground: oklch(0.985 0 0);
  --card: oklch(0.178 0.004 270);
  --muted: oklch(0.22 0.004 270);
  --muted-foreground: oklch(0.708 0.004 270);
  --border: oklch(1 0 0 / 8%);
  --input: oklch(1 0 0 / 12%);
  --primary: oklch(0.62 0.18 275);
  --primary-foreground: oklch(0.15 0.02 275);
  --primary-hover: oklch(0.57 0.19 275);
  --success: oklch(0.65 0.15 160);
  --warning: oklch(0.75 0.15 75);
  --danger: oklch(0.65 0.2 25);
}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{font-family:-apple-system,BlinkMacSystemFont,"Helvetica Neue","PingFang SC","Microsoft YaHei",sans-serif;background:var(--background);color:var(--foreground);line-height:1.6;font-size:14px;min-height:100vh;-webkit-font-smoothing:antialiased;font-feature-settings:"tnum"}
::selection{background:oklch(0.62 0.18 275 / 20%);color:var(--foreground)}
.nav{border-bottom:1px solid var(--border);background:var(--card)}
.nav-inner{max-width:1040px;margin:0 auto;padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:52px}
.nav-brand{display:flex;align-items:center;gap:8px;font-size:13.5px;font-weight:600;color:var(--foreground);text-decoration:none;letter-spacing:-.01em}
.nav-brand-mark{width:20px;height:20px;border-radius:5px;background:var(--muted);border:1px solid var(--border);display:flex;align-items:center;justify-content:center;color:var(--muted-foreground);font-size:11px;font-weight:700}
.nav-brand-sub{color:var(--muted-foreground);font-weight:400;font-size:11.5px;margin-left:2px}
.nav-links{display:flex;gap:1px;align-items:center}
.nav-links a{padding:6px 11px;border-radius:6px;font-size:12.5px;font-weight:500;color:var(--muted-foreground);text-decoration:none;transition:color .15s,background .15s}
.nav-links a:hover{color:var(--foreground);background:var(--muted)}
.nav-links a.active{color:var(--foreground);background:var(--muted)}
.wrap{max-width:1040px;margin:0 auto;padding:32px 24px 64px}
.page-head{margin-bottom:24px}
.page-head h1{font-size:20px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;margin-bottom:3px}
.page-head p{font-size:12.5px;color:var(--muted-foreground)}
.card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:22px;margin-bottom:14px}
.card h2{font-size:13px;font-weight:600;color:var(--foreground);margin-bottom:14px;letter-spacing:-.01em}
.btn{display:inline-flex;align-items:center;gap:6px;padding:7px 15px;border-radius:6px;font-size:12.5px;font-weight:500;text-decoration:none;transition:background .15s,border-color .15s,color .15s;cursor:pointer;border:1px solid transparent;font-family:inherit;line-height:1.4}
.btn-primary{background:var(--primary);color:var(--primary-foreground);font-weight:600}
.btn-primary:hover:not(:disabled){background:var(--primary-hover)}
.btn-primary:disabled{opacity:.4;cursor:not-allowed}
.btn-ghost{background:transparent;color:var(--muted-foreground);border:1px solid var(--border)}
.btn-ghost:hover{color:var(--foreground);border-color:var(--input);background:var(--muted)}
.toolbar{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.toolbar-left{display:flex;gap:8px;align-items:center}
.select{padding:6px 10px;border:1px solid var(--input);border-radius:6px;font-size:12.5px;background:var(--background);color:var(--foreground);font-family:inherit;outline:none}
.select:focus{border-color:var(--primary);box-shadow:0 0 0 3px oklch(0.62 0.18 275 / 15%)}
.stat-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:1px;background:var(--border);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:14px}
.stat-cell{background:var(--card);padding:14px 18px;display:flex;flex-direction:column;gap:3px}
.stat-label{font-size:11px;color:var(--muted-foreground);letter-spacing:.02em}
.stat-value{font-size:20px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;font-feature-settings:"tnum"}
.report-table{width:100%;border-collapse:collapse;font-size:12.5px}
.report-table th{background:var(--muted);color:var(--muted-foreground);padding:9px 12px;text-align:left;font-weight:500;border-bottom:1px solid var(--border);font-size:11px;letter-spacing:.04em;text-transform:uppercase}
.report-table td{padding:10px 12px;border-bottom:1px solid var(--border);color:var(--foreground);font-feature-settings:"tnum";vertical-align:middle}
.report-table tr:last-child td{border-bottom:none}
.report-table tr.l1-row td{background:var(--muted);font-weight:600}
.report-table .amount{text-align:right}
.bar-container{width:80px;height:5px;background:var(--input);border-radius:3px;overflow:hidden;display:inline-block;vertical-align:middle}
.bar-fill{height:100%;background:var(--primary);border-radius:3px}
.recent-table{width:100%;border-collapse:collapse;font-size:12px}
.recent-table th{background:var(--muted);color:var(--muted-foreground);padding:8px 10px;text-align:left;font-weight:500;border-bottom:1px solid var(--border);font-size:11px;letter-spacing:.04em;text-transform:uppercase}
.recent-table td{padding:9px 10px;border-bottom:1px solid var(--border);color:var(--foreground);font-feature-settings:"tnum"}
.recent-table tr:last-child td{border-bottom:none}
.recent-table tr:hover td{background:var(--muted)}
.conf-high{color:var(--success)}
.conf-mid{color:var(--warning)}
.conf-low{color:var(--danger)}
.ai-box{background:var(--card);border:1px solid var(--border);border-left:2px solid var(--primary);border-radius:0 8px 8px 0;padding:14px 18px}
.ai-box p{font-size:12.5px;color:var(--muted-foreground);line-height:1.7;margin:6px 0}
.ai-box .tag{display:inline-block;background:oklch(0.62 0.18 275 / 12%);color:var(--primary);font-size:10.5px;padding:1px 7px;border-radius:4px;margin-right:6px;font-weight:600}
.ai-loading{color:var(--muted-foreground);font-size:12.5px;padding:8px 0}
.ai-loading .spin{display:inline-block;width:14px;height:14px;border:2px solid var(--border);border-top-color:var(--primary);border-radius:50%;animation:spin .8s linear infinite;margin-right:8px;vertical-align:middle}
@keyframes spin{to{transform:rotate(360deg)}}
.feishu-row{display:flex;gap:8px;align-items:center;margin-top:12px;flex-wrap:wrap}
.chat-select{padding:7px 11px;border:1px solid var(--input);border-radius:6px;font-size:12.5px;background:var(--background);color:var(--foreground);font-family:inherit;outline:none;min-width:200px}
.result-msg{margin-top:10px;padding:9px 12px;border-radius:6px;font-size:12px;display:none}
.result-msg.success{background:oklch(0.65 0.15 160 / 10%);color:var(--success);border:1px solid oklch(0.65 0.15 160 / 25%);display:block}
.result-msg.error{background:oklch(0.65 0.2 25 / 10%);color:var(--danger);border:1px solid oklch(0.65 0.2 25 / 25%);display:block}
.empty{text-align:center;padding:32px;color:var(--muted-foreground);font-size:13px}
</style>
</head>
<body>
<nav class="nav"><div class="nav-inner">
    <a class="nav-brand" href="/"><span class="nav-brand-mark">F</span><span>财务 Agent<span class="nav-brand-sub">/ Finance</span></span></a>
    <div class="nav-links">
        <a href="/">首页</a>
        <a href="/upload">上传发票</a>
        <a href="/report" class="active">管报</a>
        <a href="/performance">绩效</a>
        <a href="/monitor">监控</a>
        <a href="/logout">退出</a>
    </div>
</div></nav>

<div class="wrap">
  <div class="page-head">
    <h1>管理报表</h1>
    <p>按收支科目层级汇总,自动生成 AI 简评,可推送飞书群</p>
  </div>

  <div class="toolbar">
    <div class="toolbar-left">
      <select class="select" id="scope-select" onchange="loadReport()">
        <option value="mine">只看我的</option>
        <option value="all">全部(仅财务)</option>
      </select>
    </div>
    <button class="btn btn-ghost" onclick="loadReport()">刷新</button>
  </div>

  <div class="stat-grid" id="stat-grid">
    <div class="stat-cell"><div class="stat-label">总笔数</div><div class="stat-value" id="total-count">—</div></div>
    <div class="stat-cell"><div class="stat-label">总金额</div><div class="stat-value" id="total-amount">—</div></div>
    <div class="stat-cell"><div class="stat-label">科目数</div><div class="stat-value" id="cat-count">—</div></div>
  </div>

  <div class="card">
    <h2>科目汇总</h2>
    <table class="report-table">
      <thead><tr><th>一级科目</th><th>二级科目</th><th class="amount">笔数</th><th class="amount">金额</th><th class="amount">占比</th></tr></thead>
      <tbody id="report-tbody"><tr><td colspan="5" class="empty">加载中...</td></tr></tbody>
    </table>
  </div>

  <div class="card">
    <h2>最新上传记录 <span style="color:var(--muted-foreground);font-weight:400;font-size:11px;margin-left:6px">最近 10 条 · scope: <span id="scope-label">—</span></span></h2>
    <table class="recent-table">
      <thead><tr><th>时间</th><th>来源</th><th>销售方/摘要</th><th>发票号</th><th class="amount">金额</th><th>科目</th><th>置信度</th><th>归属</th></tr></thead>
      <tbody id="recent-tbody"><tr><td colspan="8" class="empty">加载中...</td></tr></tbody>
    </table>
  </div>

  <div class="card">
    <h2>AI 简评</h2>
    <div id="commentary-area"><div class="ai-loading"><span class="spin"></span>AI 正在分析管报...</div></div>
  </div>

  <div class="card">
    <h2>发送到飞书</h2>
    <p style="font-size:12px;color:var(--muted-foreground);margin-bottom:10px">选择群聊后,把管报 + AI 简评发到飞书群</p>
    <div class="feishu-row">
      <select class="chat-select" id="chat-select"><option value="">加载群聊中...</option></select>
      <button class="btn btn-primary" id="send-feishu" disabled>发送到飞书</button>
    </div>
    <div class="result-msg" id="feishu-result"></div>
  </div>
</div>

<script>
function escapeHtml(s){if(s==null)return'';return String(s).replace(/[&<>"']/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]))}
async function loadReport(){
  const scope = document.getElementById('scope-select').value;
  document.getElementById('report-tbody').innerHTML = '<tr><td colspan="5" class="empty">加载中...</td></tr>';
  document.getElementById('recent-tbody').innerHTML = '<tr><td colspan="8" class="empty">加载中...</td></tr>';
  try{
    const r = await fetch('/api/report/preview?scope='+scope);
    const d = await r.json();
    if(!d.ok){ document.getElementById('report-tbody').innerHTML = '<tr><td colspan="5" class="empty">'+escapeHtml(d.error||'加载失败')+'</td></tr>'; return; }
    document.getElementById('total-count').textContent = d.total_count+' 笔';
    document.getElementById('total-amount').textContent = '¥'+(d.total_amount||0).toLocaleString();
    document.getElementById('cat-count').textContent = Object.keys(d.summary).length+' 个';
    document.getElementById('scope-label').textContent = d.scope||'?';
    // 科目汇总表
    const tbody = document.getElementById('report-tbody');
    const gt = d.total_amount || 1;
    let html = '';
    for(const [l1, info] of Object.entries(d.summary)){
      html += '<tr class="l1-row"><td>'+escapeHtml(l1)+'</td><td>—</td><td class="amount">'+info.count+'</td><td class="amount">¥'+(info.total||0).toLocaleString()+'</td><td class="amount">'+(((info.total||0)/gt)*100).toFixed(1)+'%</td></tr>';
      for(const item of (info.items||[])){
        const ratio = info.total>0?(item.total/info.total*100):0;
        html += '<tr><td>└</td><td>'+escapeHtml(item.level2)+'</td><td class="amount">'+item.count+'</td><td class="amount">¥'+(item.total||0).toLocaleString()+'</td><td class="amount"><div class="bar-container"><div class="bar-fill" style="width:'+ratio+'%"></div></div> '+ratio.toFixed(0)+'%</td></tr>';
      }
    }
    tbody.innerHTML = html || '<tr><td colspan="5" class="empty">暂无数据</td></tr>';
    // 最近记录
    const rbody = document.getElementById('recent-tbody');
    if(d.recent && d.recent.length){
      rbody.innerHTML = d.recent.map(r=>{
        const dt = r.created_at ? r.created_at.slice(5,16) : '-';
        const conf = r.confidence>=0.8?'<span class="conf-high">高</span>':r.confidence>=0.5?'<span class="conf-mid">中</span>':'<span class="conf-low">低</span>';
        const vendor = r.vendor || r.summary || '-';
        const inv = r.invoice_no || '-';
        return '<tr><td>'+escapeHtml(dt)+'</td><td>'+escapeHtml(r.source)+'</td><td>'+escapeHtml(vendor)+'</td><td>'+escapeHtml(inv)+'</td><td class="amount">¥'+(r.amount||0).toLocaleString()+'</td><td>'+escapeHtml(r.level1)+'/'+escapeHtml(r.level2||'')+'</td><td>'+conf+' '+(r.confidence*100).toFixed(0)+'%</td><td>'+escapeHtml(r.user_name)+'</td></tr>';
      }).join('');
    } else {
      rbody.innerHTML = '<tr><td colspan="8" class="empty">暂无记录</td></tr>';
    }
  }catch(e){
    document.getElementById('report-tbody').innerHTML = '<tr><td colspan="5" class="empty">网络错误:'+escapeHtml(e.message)+'</td></tr>';
  }
}
async function loadCommentary(){
  try{
    const r = await fetch('/api/report/commentary',{method:'POST'});
    const d = await r.json();
    const area = document.getElementById('commentary-area');
    if(d.ok && d.commentary){
      area.innerHTML = '<div class="ai-box">'+d.commentary.split('\\n').map(p=>p.trim()?'<p><span class="tag">AI</span>'+escapeHtml(p)+'</p>':'').join('')+'</div>';
    } else {
      area.innerHTML = '<p style="color:var(--muted-foreground);font-size:12.5px">'+escapeHtml(d.error||'简评生成失败')+'</p>';
    }
  }catch(e){
    document.getElementById('commentary-area').innerHTML = '<p style="color:var(--danger);font-size:12.5px">网络错误:'+escapeHtml(e.message)+'</p>';
  }
}
async function loadChats(){
  try{
    const r = await fetch('/api/feishu/chats');
    const d = await r.json();
    const sel = document.getElementById('chat-select');
    if(d.ok && d.chats && d.chats.length>0){
      sel.innerHTML = d.chats.map(c=>'<option value="'+escapeHtml(c.chat_id)+'">'+escapeHtml(c.name)+'</option>').join('');
      document.getElementById('send-feishu').disabled = false;
    } else {
      sel.innerHTML = '<option value="">无可用群聊 ('+escapeHtml(d.error||'未知错误')+')</option>';
    }
  }catch(e){
    document.getElementById('chat-select').innerHTML = '<option value="">加载失败 ('+escapeHtml(e.message)+')</option>';
  }
}
document.getElementById('send-feishu').addEventListener('click', async ()=>{
  const chatId = document.getElementById('chat-select').value;
  if(!chatId) return;
  const btn = document.getElementById('send-feishu');
  const msg = document.getElementById('feishu-result');
  btn.disabled = true; btn.textContent = '发送中...';
  msg.className = 'result-msg';
  try{
    const r = await fetch('/api/report/feishu',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({chat_id:chatId})});
    const d = await r.json();
    if(d.ok){ msg.className='result-msg success'; msg.textContent='已发送到飞书群'; }
    else{ msg.className='result-msg error'; msg.textContent='失败:'+(d.error||'未知错误'); }
  }catch(e){ msg.className='result-msg error'; msg.textContent='网络错误:'+e.message; }
  btn.disabled = false; btn.textContent = '发送到飞书';
});
loadReport();
loadCommentary();
loadChats();
</script>
</body>
</html>"""


PERFORMANCE_HTML = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>绩效试算 · 财务 Agent</title>
<style>
:root{
  --background: oklch(0.145 0.004 270);
  --foreground: oklch(0.985 0 0);
  --card: oklch(0.178 0.004 270);
  --card-foreground: oklch(0.985 0 0);
  --muted: oklch(0.22 0.004 270);
  --muted-foreground: oklch(0.708 0.004 270);
  --border: oklch(1 0 0 / 8%);
  --input: oklch(1 0 0 / 12%);
  --ring: oklch(0.62 0.18 275);
  --primary: oklch(0.62 0.18 275);
  --primary-foreground: oklch(0.15 0.02 275);
  --primary-hover: oklch(0.57 0.19 275);
  --success: oklch(0.65 0.15 160);
  --warning: oklch(0.75 0.15 75);
  --danger: oklch(0.65 0.2 25);
  --radius: 0.375rem;
}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{
  font-family:-apple-system,BlinkMacSystemFont,"Helvetica Neue","PingFang SC","Microsoft YaHei",sans-serif;
  background:var(--background);color:var(--foreground);line-height:1.6;font-size:14px;min-height:100vh;
  -webkit-font-smoothing:antialiased;font-feature-settings:"tnum";
}
::selection{background:oklch(0.62 0.18 275 / 20%);color:var(--foreground)}
.nav{border-bottom:1px solid var(--border);background:var(--card)}
.nav-inner{max-width:1040px;margin:0 auto;padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:52px}
.nav-brand{display:flex;align-items:center;gap:8px;font-size:13.5px;font-weight:600;color:var(--foreground);text-decoration:none;letter-spacing:-.01em}
.nav-brand-mark{width:20px;height:20px;border-radius:5px;background:var(--muted);border:1px solid var(--border);display:flex;align-items:center;justify-content:center;color:var(--muted-foreground);font-size:11px;font-weight:700}
.nav-brand-sub{color:var(--muted-foreground);font-weight:400;font-size:11.5px;margin-left:2px}
.nav-links{display:flex;gap:1px;align-items:center}
.nav-links a{padding:6px 11px;border-radius:6px;font-size:12.5px;font-weight:500;color:var(--muted-foreground);text-decoration:none;transition:color .15s,background .15s}
.nav-links a:hover{color:var(--foreground);background:var(--muted)}
.nav-links a.active{color:var(--foreground);background:var(--muted)}
.wrap{max-width:1040px;margin:0 auto;padding:32px 24px 64px}
.page-head{margin-bottom:24px}
.page-head h1{font-size:20px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;margin-bottom:3px}
.page-head p{font-size:12.5px;color:var(--muted-foreground)}
.card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:22px;margin-bottom:14px}
.card h2{font-size:13px;font-weight:600;color:var(--foreground);margin-bottom:3px;letter-spacing:-.01em}
.muted{color:var(--muted-foreground);font-size:12px}
.btn{display:inline-flex;align-items:center;gap:6px;padding:7px 15px;border-radius:6px;font-size:12.5px;font-weight:500;text-decoration:none;transition:background .15s,border-color .15s,color .15s;cursor:pointer;border:1px solid transparent;font-family:inherit;line-height:1.4}
.btn-primary{background:var(--primary);color:var(--primary-foreground);font-weight:600}
.btn-primary:hover{background:var(--primary-hover)}
.btn-ghost{background:transparent;color:var(--muted-foreground);border:1px solid var(--border)}
.btn-ghost:hover{color:var(--foreground);border-color:var(--input);background:var(--muted)}

.toolbar{display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:16px;flex-wrap:wrap}
.toolbar-left{display:flex;gap:8px;align-items:center}
.select{padding:6px 10px;border:1px solid var(--input);border-radius:6px;font-size:12.5px;background:var(--background);color:var(--foreground);font-family:inherit;outline:none}
.select:focus{border-color:var(--primary);box-shadow:0 0 0 3px oklch(0.62 0.18 275 / 15%)}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:1px;background:var(--border);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:16px}
.kpi-card{background:var(--card);padding:14px 16px;display:flex;flex-direction:column;gap:3px}
.kpi-label{font-size:11px;color:var(--muted-foreground);letter-spacing:.02em}
.kpi-value{font-size:20px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;font-feature-settings:"tnum"}
.kpi-value .unit{font-size:11px;color:var(--muted-foreground);font-weight:400;margin-left:2px}
.perf-table{width:100%;border-collapse:collapse;font-size:12.5px;background:var(--card);border:1px solid var(--border);border-radius:8px;overflow:hidden}
.perf-table th{background:var(--muted);color:var(--muted-foreground);padding:10px 14px;text-align:left;font-weight:500;border-bottom:1px solid var(--border);font-size:11px;letter-spacing:.04em;text-transform:uppercase}
.perf-table td{padding:11px 14px;border-bottom:1px solid var(--border);color:var(--foreground);font-feature-settings:"tnum"}
.perf-table tr:last-child td{border-bottom:none}
.perf-table tr:hover td{background:var(--muted)}
.perf-table .amount{text-align:right}
.perf-table .total-row td{background:var(--muted);font-weight:600;border-top:1px solid var(--input)}
.empty{text-align:center;padding:40px;color:var(--muted-foreground);font-size:13px}
</style>
</head>
<body>
<nav class="nav"><div class="nav-inner">
    <a class="nav-brand" href="/"><span class="nav-brand-mark">F</span><span>财务 Agent<span class="nav-brand-sub">/ Finance</span></span></a>
    <div class="nav-links">
        <a href="/">首页</a>
        <a href="/upload">上传发票</a>
        <a href="/report">管报</a>
        <a href="/performance">绩效</a>
        <a href="/monitor">监控</a>
        <a href="/logout">退出</a>
    </div>
</div></nav>

<div class="wrap">
  <div class="page-head">
    <h1>绩效试算</h1>
    <p>按部门/项目维度汇总成本与 KPI</p>
  </div>

  <div class="toolbar">
    <div class="toolbar-left">
      <select class="select" id="dept-select" onchange="loadPerf()">
        <option value="">全部部门</option>
      </select>
    </div>
    <button class="btn btn-ghost" onclick="loadPerf()">刷新</button>
  </div>

  <div id="kpi-grid" class="kpi-grid"></div>
  <div id="perf-content">
    <div class="empty">加载中...</div>
  </div>
</div>

<script>
async function loadPerf(){
  const dept = document.getElementById('dept-select').value;
  const container = document.getElementById('perf-content');
  const kpi = document.getElementById('kpi-grid');
  container.innerHTML = '<div class="empty">加载中...</div>';
  kpi.innerHTML = '';
  try{
    const url = '/api/performance/calculate'+(dept?('?department='+encodeURIComponent(dept)):'');
    const r = await fetch(url);
    const d = await r.json();
    if(!d.ok){ container.innerHTML = '<div class="empty">'+(d.error||'未配置绩效规则或加载失败')+'</div>'; return; }
    const depts = d.departments || [];
    // KPI 卡片
    const kpis = [
      {label:'部门数', value:depts.length, unit:''},
      {label:'累计流水', value:(d.total_amount||0).toLocaleString(), unit:'元'},
      {label:'累计笔数', value:d.total_count||0, unit:'笔'},
      {label:'应发奖金', value:(d.total_bonus||0).toLocaleString(), unit:'元'},
    ];
    kpi.innerHTML = kpis.map(k=>'<div class="kpi-card"><div class="kpi-label">'+k.label+'</div><div class="kpi-value">'+k.value+'<span class="unit">'+k.unit+'</span></div></div>').join('');
    if(depts.length === 0){ container.innerHTML = '<div class="empty">暂无数据</div>'; return; }
    // 填充部门筛选
    const sel = document.getElementById('dept-select');
    if(sel.options.length <= 1){
      depts.forEach(r=>{
        const o = document.createElement('option');
        o.value = r.department; o.textContent = r.department;
        sel.appendChild(o);
      });
    }
    let html = '<table class="perf-table"><thead><tr><th>部门</th><th>岗位</th><th class="amount">目标</th><th class="amount">实际</th><th class="amount">达成率</th><th class="amount">奖金</th><th>状态</th></tr></thead><tbody>';
    depts.forEach(r=>{
      html += '<tr><td>'+r.department+'</td><td>'+(r.position||'')+'</td><td class="amount">'+(r.target_amount||0).toLocaleString()+'</td><td class="amount">'+(r.actual_amount||0).toLocaleString()+'</td><td class="amount">'+(r.achievement_rate||0)+'%</td><td class="amount">'+(r.bonus||0).toLocaleString()+'</td><td>'+(r.status||'')+'</td></tr>';
    });
    html += '<tr class="total-row"><td>合计</td><td></td><td></td><td></td><td></td><td class="amount">'+(d.total_bonus||0).toLocaleString()+'</td><td></td></tr>';
    html += '</tbody></table>';
    container.innerHTML = html;
  }catch(e){
    container.innerHTML = '<div class="empty">网络错误:'+e.message+'</div>';
  }
}
loadPerf();
</script>
</body>
</html>"""

INDEX_HTML = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>财务 Agent · 发票试算与管报系统</title>
<style>
:root{
  --background: oklch(0.145 0.004 270);
  --foreground: oklch(0.985 0 0);
  --card: oklch(0.178 0.004 270);
  --card-foreground: oklch(0.985 0 0);
  --muted: oklch(0.22 0.004 270);
  --muted-foreground: oklch(0.708 0.004 270);
  --border: oklch(1 0 0 / 8%);
  --input: oklch(1 0 0 / 12%);
  --ring: oklch(0.62 0.18 275);
  --primary: oklch(0.62 0.18 275);
  --primary-foreground: oklch(0.15 0.02 275);
  --primary-hover: oklch(0.57 0.19 275);
  --success: oklch(0.65 0.15 160);
  --warning: oklch(0.75 0.15 75);
  --danger: oklch(0.65 0.2 25);
  --radius: 0.375rem;
}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{font-family:-apple-system,BlinkMacSystemFont,"Helvetica Neue","PingFang SC","Microsoft YaHei",sans-serif;background:var(--background);color:var(--foreground);line-height:1.6;font-size:14px;min-height:100vh;-webkit-font-smoothing:antialiased;font-feature-settings:"tnum"}
::selection{background:oklch(0.62 0.18 275 / 20%);color:var(--foreground)}
.nav{border-bottom:1px solid var(--border);background:var(--card)}
.nav-inner{max-width:1040px;margin:0 auto;padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:52px}
.nav-brand{display:flex;align-items:center;gap:8px;font-size:13.5px;font-weight:600;color:var(--foreground);text-decoration:none;letter-spacing:-.01em}
.nav-brand-mark{width:20px;height:20px;border-radius:5px;background:var(--muted);border:1px solid var(--border);display:flex;align-items:center;justify-content:center;color:var(--muted-foreground);font-size:11px;font-weight:700}
.nav-brand-sub{color:var(--muted-foreground);font-weight:400;font-size:11.5px;margin-left:2px}
.nav-links{display:flex;gap:1px;align-items:center}
.nav-links a{padding:6px 11px;border-radius:6px;font-size:12.5px;font-weight:500;color:var(--muted-foreground);text-decoration:none;transition:color .15s,background .15s}
.nav-links a:hover{color:var(--foreground);background:var(--muted)}
.nav-links a.active{color:var(--foreground);background:var(--muted)}
.wrap{max-width:1040px;margin:0 auto;padding:32px 24px 64px}
.hero{margin-bottom:48px;max-width:680px}
.hero-eyebrow{font-size:11.5px;color:var(--muted-foreground);font-weight:500;letter-spacing:.04em;margin-bottom:16px;text-transform:uppercase}
.hero h1{font-size:28px;font-weight:600;letter-spacing:-.02em;line-height:1.3;margin-bottom:14px;color:var(--foreground)}
.hero h1 .accent{color:var(--primary)}
.hero p{font-size:14px;color:var(--muted-foreground);line-height:1.7;margin-bottom:24px;max-width:560px}
.hero-actions{display:flex;gap:8px;flex-wrap:wrap}
.btn{display:inline-flex;align-items:center;gap:6px;padding:7px 15px;border-radius:6px;font-size:12.5px;font-weight:500;text-decoration:none;transition:background .15s,border-color .15s,color .15s;cursor:pointer;border:1px solid transparent;font-family:inherit;line-height:1.4}
.btn-primary{background:var(--primary);color:var(--primary-foreground);font-weight:600}
.btn-primary:hover{background:var(--primary-hover)}
.btn-ghost{background:transparent;color:var(--muted-foreground);border:1px solid var(--border)}
.btn-ghost:hover{color:var(--foreground);border-color:var(--input);background:var(--muted)}
.metrics{display:grid;grid-template-columns:repeat(4,1fr);gap:1px;background:var(--border);border:1px solid var(--border);border-radius:8px;overflow:hidden;margin-bottom:48px}
.metric{background:var(--card);padding:16px 18px;display:flex;flex-direction:column;gap:4px}
.metric-value{font-size:22px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;font-feature-settings:"tnum"}
.metric-value .unit{font-size:12px;color:var(--muted-foreground);font-weight:400;margin-left:2px}
.metric-label{font-size:11px;color:var(--muted-foreground);letter-spacing:.02em}
.metric-status{display:inline-flex;align-items:center;gap:6px;font-size:13px;color:var(--success);font-weight:500}
.metric-status::before{content:'';width:6px;height:6px;border-radius:50%;background:var(--success)}
.section{margin-bottom:48px}
.section-head{display:flex;align-items:baseline;justify-content:space-between;margin-bottom:16px;padding-bottom:10px;border-bottom:1px solid var(--border)}
.section-title{font-size:13.5px;font-weight:600;color:var(--foreground);letter-spacing:-.01em}
.section-sub{font-size:11.5px;color:var(--muted-foreground)}
.feat-list{display:flex;flex-direction:column}
.feat{display:grid;grid-template-columns:32px 1fr auto;gap:14px;padding:14px 0;border-bottom:1px solid var(--border);align-items:start;transition:background .15s}
.feat:last-child{border-bottom:none}
.feat:hover{background:var(--muted)}
.feat-num{font-size:11px;color:var(--muted-foreground);font-weight:500;font-feature-settings:"tnum";padding-top:3px;letter-spacing:.05em}
.feat-body h3{font-size:13.5px;font-weight:600;color:var(--foreground);margin-bottom:3px;letter-spacing:-.01em}
.feat-body p{font-size:12.5px;color:var(--muted-foreground);line-height:1.6}
.feat-status{font-size:10.5px;color:var(--success);font-weight:500;padding:2px 7px;border:1px solid oklch(0.65 0.15 160 / 25%);border-radius:4px;white-space:nowrap;background:oklch(0.65 0.15 160 / 8%)}
.entries{display:grid;grid-template-columns:1.4fr 1fr 1fr;gap:10px;margin-bottom:48px}
.entry{display:block;padding:18px;background:var(--card);border:1px solid var(--border);border-radius:8px;text-decoration:none;transition:border-color .15s,background .15s}
.entry:hover{border-color:var(--input);background:var(--muted)}
.entry-title{font-size:13.5px;font-weight:600;color:var(--foreground);margin-bottom:4px;display:flex;align-items:center;justify-content:space-between}
.entry-arrow{color:var(--muted-foreground);font-size:14px;transition:color .15s}
.entry:hover .entry-arrow{color:var(--primary)}
.entry-desc{font-size:12px;color:var(--muted-foreground);line-height:1.5}
.footer{padding-top:24px;border-top:1px solid var(--border);font-size:11px;color:var(--muted-foreground);display:flex;justify-content:space-between;flex-wrap:wrap;gap:8px}
@media(max-width:720px){.nav-links{display:none}.hero h1{font-size:22px}.metrics{grid-template-columns:1fr 1fr}.entries{grid-template-columns:1fr}.feat{grid-template-columns:24px 1fr}.feat-status{grid-column:2}}
</style>
</head>
<body>
<nav class="nav"><div class="nav-inner">
    <a class="nav-brand" href="/"><span class="nav-brand-mark">F</span><span>财务 Agent<span class="nav-brand-sub">/ Finance</span></span></a>
    <div class="nav-links">
        <a href="/">首页</a>
        <a href="/upload">上传发票</a>
        <a href="/report">管报</a>
        <a href="/performance">绩效</a>
        <a href="/monitor">监控</a>
        <a href="/logout">退出</a>
    </div>
</div></nav>

<div class="wrap">
  <section class="hero">
    <div class="hero-eyebrow">Finance Agent · v1.0</div>
    <h1>发票试算与<span class="accent">管理报表</span>智能生成</h1>
    <p>上传发票 PDF/图片/TXT,AI 一步提取并归一化(科目 + 层级 + 置信度),自动生成管报与绩效试算,支持飞书 Bot 上传与企业级权限分层。</p>
    <div class="hero-actions">
      <a class="btn btn-primary" href="/upload">上传发票 →</a>
      <a class="btn btn-ghost" href="/report">查看管报</a>
      <a class="btn btn-ghost" href="/performance">绩效试算</a>
    </div>
  </section>

  <div class="metrics">
    <div class="metric"><div class="metric-value"><span id="s-tx">—</span><span class="unit">笔</span></div><div class="metric-label">累计发票</div></div>
    <div class="metric"><div class="metric-value"><span id="s-today">—</span><span class="unit">笔</span></div><div class="metric-label">今日上传</div></div>
    <div class="metric"><div class="metric-value"><span id="s-amount">—</span><span class="unit">元</span></div><div class="metric-label">累计金额</div></div>
    <div class="metric"><div class="metric-value metric-status">在线</div><div class="metric-label" id="s-uptime">服务状态</div></div>
  </div>

  <div class="section">
    <div class="section-head">
      <div class="section-title">功能模块</div>
      <div class="section-sub">8 天交付 · 49 个单元测试全绿</div>
    </div>
    <div class="feat-list">
      <div class="feat"><div class="feat-num">01</div><div class="feat-body"><h3>发票 OCR + AI 解析</h3><p>pypdf 文本提取 + tesseract OCR 降级,AI 一步提取发票号/日期/供应商/金额/明细,并归一化到科目层级。</p></div><div class="feat-status">已上线</div></div>
      <div class="feat"><div class="feat-num">02</div><div class="feat-body"><h3>管报生成</h3><p>按收支科目汇总,一级 + 二级层级展示,支持 scope=all(财务)/scope=mine(员工)权限分层。</p></div><div class="feat-status">已上线</div></div>
      <div class="feat"><div class="feat-num">03</div><div class="feat-body"><h3>绩效试算</h3><p>按部门/项目维度汇总成本,支持自定义 KPI 与权重,输出绩效试算表。</p></div><div class="feat-status">已上线</div></div>
      <div class="feat"><div class="feat-num">04</div><div class="feat-body"><h3>飞书 Bot 上传</h3><p>飞书对话直接发送图片/文件,自动解析入库,open_id 自动建档为 employee 角色。</p></div><div class="feat-status">已上线</div></div>
      <div class="feat"><div class="feat-num">05</div><div class="feat-body"><h3>权限分层</h3><p>financial 角色看全部数据,employee 只看自己上传的,登录态用 Flask session + secret_key。</p></div><div class="feat-status">已上线</div></div>
      <div class="feat"><div class="feat-num">06</div><div class="feat-body"><h3>监控告警</h3><p>异常自动告警,支持飞书群推送(配置 ALERT_CHAT_ID 后启用),Dashboard 实时展示服务状态。</p></div><div class="feat-status">已上线</div></div>
    </div>
  </div>

  <div class="entries">
    <a class="entry" href="/upload"><div class="entry-title">上传发票开始试算 <span class="entry-arrow">→</span></div><div class="entry-desc">PDF/图片/TXT,OCR + AI 归一化,约 10-30 秒出结果。</div></a>
    <a class="entry" href="/report"><div class="entry-title">管理报表 <span class="entry-arrow">→</span></div><div class="entry-desc">收支汇总 + 层级展示。</div></a>
    <a class="entry" href="/performance"><div class="entry-title">绩效试算 <span class="entry-arrow">→</span></div><div class="entry-desc">按部门/项目维度。</div></a>
  </div>

  <div class="footer">
    <span>财务 Agent · 企业运营智能化 Demo</span>
    <span>基于 2026-07-09 线下拜访会议需求 · 8 天敏捷交付</span>
  </div>
</div>

<script>
async function loadStats(){
  try{
    const r = await fetch('/api/stats').then(r=>r.json()).catch(()=>null);
    if(r){
      document.getElementById('s-tx').textContent = r.total_transactions ?? '—';
      document.getElementById('s-today').textContent = r.today_transactions ?? '—';
      document.getElementById('s-amount').textContent = r.total_amount ?? '—';
      if(r.uptime_human) document.getElementById('s-uptime').textContent = '运行 ' + r.uptime_human;
    }
  }catch(e){}
}
loadStats();
</script>
</body>
</html>"""

LOGIN_HTML = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>登录 · 财务 Agent</title>
<style>
:root{
  --background: oklch(0.145 0.004 270);
  --foreground: oklch(0.985 0 0);
  --card: oklch(0.178 0.004 270);
  --card-foreground: oklch(0.985 0 0);
  --muted: oklch(0.22 0.004 270);
  --muted-foreground: oklch(0.708 0.004 270);
  --border: oklch(1 0 0 / 8%);
  --input: oklch(1 0 0 / 12%);
  --ring: oklch(0.62 0.18 275);
  --primary: oklch(0.62 0.18 275);
  --primary-foreground: oklch(0.15 0.02 275);
  --primary-hover: oklch(0.57 0.19 275);
  --success: oklch(0.65 0.15 160);
  --warning: oklch(0.75 0.15 75);
  --danger: oklch(0.65 0.2 25);
  --radius: 0.375rem;
}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{
  font-family:-apple-system,BlinkMacSystemFont,"Helvetica Neue","PingFang SC","Microsoft YaHei",sans-serif;
  background:var(--background);color:var(--foreground);line-height:1.6;font-size:14px;min-height:100vh;
  -webkit-font-smoothing:antialiased;font-feature-settings:"tnum";
}
::selection{background:oklch(0.62 0.18 275 / 20%);color:var(--foreground)}
.nav{border-bottom:1px solid var(--border);background:var(--card)}
.nav-inner{max-width:1040px;margin:0 auto;padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:52px}
.nav-brand{display:flex;align-items:center;gap:8px;font-size:13.5px;font-weight:600;color:var(--foreground);text-decoration:none;letter-spacing:-.01em}
.nav-brand-mark{width:20px;height:20px;border-radius:5px;background:var(--muted);border:1px solid var(--border);display:flex;align-items:center;justify-content:center;color:var(--muted-foreground);font-size:11px;font-weight:700}
.nav-brand-sub{color:var(--muted-foreground);font-weight:400;font-size:11.5px;margin-left:2px}
.nav-links{display:flex;gap:1px;align-items:center}
.nav-links a{padding:6px 11px;border-radius:6px;font-size:12.5px;font-weight:500;color:var(--muted-foreground);text-decoration:none;transition:color .15s,background .15s}
.nav-links a:hover{color:var(--foreground);background:var(--muted)}
.nav-links a.active{color:var(--foreground);background:var(--muted)}
.wrap{max-width:1040px;margin:0 auto;padding:32px 24px 64px}
.page-head{margin-bottom:24px}
.page-head h1{font-size:20px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;margin-bottom:3px}
.page-head p{font-size:12.5px;color:var(--muted-foreground)}
.card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:22px;margin-bottom:14px}
.card h2{font-size:13px;font-weight:600;color:var(--foreground);margin-bottom:3px;letter-spacing:-.01em}
.muted{color:var(--muted-foreground);font-size:12px}
.btn{display:inline-flex;align-items:center;gap:6px;padding:7px 15px;border-radius:6px;font-size:12.5px;font-weight:500;text-decoration:none;transition:background .15s,border-color .15s,color .15s;cursor:pointer;border:1px solid transparent;font-family:inherit;line-height:1.4}
.btn-primary{background:var(--primary);color:var(--primary-foreground);font-weight:600}
.btn-primary:hover{background:var(--primary-hover)}
.btn-ghost{background:transparent;color:var(--muted-foreground);border:1px solid var(--border)}
.btn-ghost:hover{color:var(--foreground);border-color:var(--input);background:var(--muted)}

.login-wrap{max-width:360px;margin:80px auto;padding:0 24px}
.login-card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:32px 28px}
.login-title{font-size:18px;font-weight:600;color:var(--foreground);margin-bottom:4px;letter-spacing:-.01em}
.login-sub{font-size:12px;color:var(--muted-foreground);margin-bottom:24px}
.form-group{margin-bottom:14px}
.form-group label{display:block;font-size:11.5px;font-weight:500;color:var(--muted-foreground);margin-bottom:5px;letter-spacing:.02em}
.form-group input{width:100%;padding:9px 12px;border:1px solid var(--input);border-radius:6px;font-size:13px;font-family:inherit;background:var(--background);color:var(--foreground);outline:none;transition:border-color .15s,box-shadow .15s}
.form-group input:focus{border-color:var(--primary);box-shadow:0 0 0 3px oklch(0.62 0.18 275 / 15%)}
.btn{display:inline-flex;align-items:center;justify-content:center;gap:6px;padding:9px 16px;border-radius:6px;font-size:13px;font-weight:600;cursor:pointer;border:1px solid transparent;font-family:inherit;transition:background .15s;width:100%}
.btn-primary{background:var(--primary);color:var(--primary-foreground)}
.btn-primary:hover{background:var(--primary-hover)}
.error-msg{color:var(--danger);font-size:12px;margin-top:12px;padding:8px 12px;background:oklch(0.65 0.2 25 / 8%);border:1px solid oklch(0.65 0.2 25 / 25%);border-radius:5px}
.demo-accounts{margin-top:20px;padding-top:16px;border-top:1px solid var(--border);font-size:11px;color:var(--muted-foreground);line-height:1.7}
.demo-accounts b{color:var(--foreground);font-weight:500}
</style>
</head>
<body>
<div class="login-wrap">
  <div class="login-card">
    <div class="login-title">财务 Agent</div>
    <div class="login-sub">登录以使用发票试算与管报系统</div>
    <form method="post" action="/login">
      <div class="form-group">
        <label>姓名</label>
        <input type="text" name="name" required autofocus>
      </div>
      <div class="form-group">
        <label>PIN 码</label>
        <input type="password" name="pin" required>
      </div>
      <button type="submit" class="btn btn-primary">登录</button>
      {% if error %}<div class="error-msg">{{ error }}</div>{% endif %}
    </form>
    <div class="demo-accounts">
      <b>测试账号:</b><br>
      财务管理员 / 1234(看全部)<br>
      张三 / 1111(只看自己)<br>
      李四 / 2222(只看自己)
    </div>
  </div>
</div>
</body>
</html>"""

UPLOAD_HTML = """<!DOCTYPE html>
<html lang="zh">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>上传发票 · 财务 Agent</title>
<style>
:root{
  --background: oklch(0.145 0.004 270);
  --foreground: oklch(0.985 0 0);
  --card: oklch(0.178 0.004 270);
  --muted: oklch(0.22 0.004 270);
  --muted-foreground: oklch(0.708 0.004 270);
  --border: oklch(1 0 0 / 8%);
  --input: oklch(1 0 0 / 12%);
  --primary: oklch(0.62 0.18 275);
  --primary-foreground: oklch(0.15 0.02 275);
  --primary-hover: oklch(0.57 0.19 275);
  --success: oklch(0.65 0.15 160);
  --warning: oklch(0.75 0.15 75);
  --danger: oklch(0.65 0.2 25);
}
*{margin:0;padding:0;box-sizing:border-box}
html{scroll-behavior:smooth}
body{font-family:-apple-system,BlinkMacSystemFont,"Helvetica Neue","PingFang SC","Microsoft YaHei",sans-serif;background:var(--background);color:var(--foreground);line-height:1.6;font-size:14px;min-height:100vh;-webkit-font-smoothing:antialiased;font-feature-settings:"tnum"}
.nav{border-bottom:1px solid var(--border);background:var(--card)}
.nav-inner{max-width:1040px;margin:0 auto;padding:0 24px;display:flex;align-items:center;justify-content:space-between;height:52px}
.nav-brand{display:flex;align-items:center;gap:8px;font-size:13.5px;font-weight:600;color:var(--foreground);text-decoration:none;letter-spacing:-.01em}
.nav-brand-mark{width:20px;height:20px;border-radius:5px;background:var(--muted);border:1px solid var(--border);display:flex;align-items:center;justify-content:center;color:var(--muted-foreground);font-size:11px;font-weight:700}
.nav-brand-sub{color:var(--muted-foreground);font-weight:400;font-size:11.5px;margin-left:2px}
.nav-links{display:flex;gap:1px;align-items:center}
.nav-links a{padding:6px 11px;border-radius:6px;font-size:12.5px;font-weight:500;color:var(--muted-foreground);text-decoration:none;transition:color .15s,background .15s}
.nav-links a:hover{color:var(--foreground);background:var(--muted)}
.nav-links a.active{color:var(--foreground);background:var(--muted)}
.wrap{max-width:1040px;margin:0 auto;padding:32px 24px 64px}
.page-head{margin-bottom:24px}
.page-head h1{font-size:20px;font-weight:600;color:var(--foreground);letter-spacing:-.02em;margin-bottom:3px}
.page-head p{font-size:12.5px;color:var(--muted-foreground)}
.card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:22px;margin-bottom:14px}
.btn{display:inline-flex;align-items:center;gap:6px;padding:8px 22px;border-radius:6px;font-size:12.5px;font-weight:600;cursor:pointer;border:1px solid transparent;font-family:inherit;transition:background .15s}
.btn-primary{background:var(--primary);color:var(--primary-foreground)}
.btn-primary:hover:not(:disabled){background:var(--primary-hover)}
.btn-primary:disabled{opacity:.4;cursor:not-allowed}
.btn-ghost{background:transparent;color:var(--muted-foreground);border:1px solid var(--border);font-weight:500}
.btn-ghost:hover{color:var(--foreground);border-color:var(--input);background:var(--muted)}
.drop-zone{border:1px dashed var(--input);border-radius:8px;padding:36px;text-align:center;color:var(--muted-foreground);cursor:pointer;transition:border-color .15s,background .15s;background:var(--muted)}
.drop-zone:hover{border-color:var(--primary);background:var(--card)}
.drop-zone.dragover{border-color:var(--primary);background:var(--card)}
.drop-zone.has-file{border-color:var(--success);background:oklch(0.65 0.15 160 / 8%);color:var(--success)}
.drop-icon{font-size:24px;margin-bottom:6px;opacity:.6}
.drop-text{font-size:13px;font-weight:500;margin-bottom:2px;color:var(--foreground)}
.drop-hint{font-size:11.5px;color:var(--muted-foreground)}
.info-box{background:var(--muted);border-left:2px solid var(--primary);border-radius:0 6px 6px 0;padding:11px 14px;margin-top:14px;font-size:11.5px;color:var(--muted-foreground);line-height:1.7}
.info-box b{color:var(--foreground);font-weight:600}
.action-row{margin-top:14px;display:flex;gap:8px;align-items:center}
.loading{display:inline-block;width:14px;height:14px;border:2px solid var(--border);border-top-color:var(--primary);border-radius:50%;animation:spin .8s linear infinite;margin-right:6px;vertical-align:middle}
@keyframes spin{to{transform:rotate(360deg)}}
.result-card{background:var(--card);border:1px solid var(--border);border-radius:8px;padding:18px;margin-top:14px}
.result-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;flex-wrap:wrap;gap:8px}
.result-head h3{font-size:14px;font-weight:600;color:var(--success);letter-spacing:-.01em}
.lvl-tag{display:inline-block;padding:3px 10px;border-radius:5px;font-size:11.5px;font-weight:600;border:1px solid;background:oklch(0.62 0.18 275 / 10%);color:var(--primary);border-color:oklch(0.62 0.18 275 / 25%)}
.field-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px 28px}
.field{display:flex;flex-direction:column;gap:3px}
.field-label{font-size:11px;color:var(--muted-foreground);font-weight:500;letter-spacing:.02em}
.field-value{font-size:13px;color:var(--foreground);font-weight:500;word-break:break-all}
.field-value.amount{font-size:17px;font-weight:700;color:var(--success);font-feature-settings:"tnum"}
.items-list{display:flex;flex-wrap:wrap;gap:5px;margin-top:4px}
.item-tag{padding:3px 9px;background:var(--muted);border:1px solid var(--border);border-radius:5px;font-size:11.5px;color:var(--muted-foreground)}
.confidence-wrap{margin-top:4px}
.confidence-bar{height:5px;border-radius:3px;background:var(--input);overflow:hidden;margin-bottom:4px}
.confidence-fill{height:100%;border-radius:3px}
.confidence-label{font-size:11px;color:var(--muted-foreground)}
.conf-high{color:var(--success)}
.conf-mid{color:var(--warning)}
.conf-low{color:var(--danger)}
.ocr-preview{margin-top:12px;padding:10px 12px;background:var(--background);border:1px solid var(--border);border-radius:6px;font-size:11.5px;color:var(--muted-foreground);max-height:200px;overflow-y:auto;white-space:pre-wrap;font-family:ui-monospace,Menlo,Monaco,monospace;line-height:1.5}
details summary{cursor:pointer;font-size:11.5px;color:var(--muted-foreground);user-select:none}
details summary:hover{color:var(--foreground)}
.error-box{background:oklch(0.65 0.2 25 / 8%);border:1px solid oklch(0.65 0.2 25 / 25%);color:var(--danger);padding:11px 14px;border-radius:6px;margin-top:14px;font-size:12.5px}
.next-link{display:inline-block;margin-top:14px;color:var(--primary);font-weight:500;font-size:12.5px;text-decoration:none}
.next-link:hover{text-decoration:underline}
</style>
</head>
<body>
<nav class="nav"><div class="nav-inner">
    <a class="nav-brand" href="/"><span class="nav-brand-mark">F</span><span>财务 Agent<span class="nav-brand-sub">/ Finance</span></span></a>
    <div class="nav-links">
        <a href="/">首页</a>
        <a href="/upload" class="active">上传发票</a>
        <a href="/report">管报</a>
        <a href="/performance">绩效</a>
        <a href="/monitor">监控</a>
        <a href="/logout">退出</a>
    </div>
</div></nav>

<div class="wrap">
  <div class="page-head">
    <h1>上传发票</h1>
    <p>PDF / 图片 / TXT → OCR 提取 → AI 解析字段 + 归一化科目 → 入库</p>
  </div>

  <div class="card">
    <form id="upload-form" enctype="multipart/form-data">
      <div class="drop-zone" id="drop-zone">
        <div class="drop-icon">⬆</div>
        <div class="drop-text">点击或拖拽文件上传</div>
        <div class="drop-hint">支持 PDF / JPG / PNG / TXT(扫描件自动 OCR)· 单文件 ≤ 10MB</div>
        <input type="file" id="file-input" name="file" style="display:none" accept=".pdf,.jpg,.jpeg,.png,.txt,.tif,.tiff,.bmp">
      </div>
      <div class="info-box">
        <b>处理流程:</b> OCR 文本提取 → AI 解析(发票号/日期/供应商/金额/明细)→ 科目归一化(一级 + 二级 + 置信度)→ 入库 → 管报/绩效自动更新
      </div>
      <div class="action-row">
        <button type="submit" class="btn btn-primary" id="upload-btn" disabled>开始解析</button>
        <button type="button" class="btn btn-ghost" onclick="clearResult()">清空</button>
      </div>
    </form>

    <div id="result"></div>
  </div>
</div>

<script>
function escapeHtml(s){if(s==null)return'';return String(s).replace(/[&<>"']/g,ch=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[ch]))}
const dropZone = document.getElementById('drop-zone');
const fileInput = document.getElementById('file-input');
const uploadBtn = document.getElementById('upload-btn');
const result = document.getElementById('result');
let currentFile = null;

dropZone.addEventListener('click', ()=>fileInput.click());
dropZone.addEventListener('dragover', e=>{e.preventDefault();dropZone.classList.add('dragover');});
dropZone.addEventListener('dragleave', ()=>dropZone.classList.remove('dragover'));
dropZone.addEventListener('drop', e=>{
  e.preventDefault();
  dropZone.classList.remove('dragover');
  if(e.dataTransfer.files.length){ handleFile(e.dataTransfer.files[0]); }
});
fileInput.addEventListener('change', e=>{ if(e.target.files.length) handleFile(e.target.files[0]); });

function handleFile(file){
  currentFile = file;
  dropZone.classList.add('has-file');
  dropZone.querySelector('.drop-icon').textContent = '✓';
  dropZone.querySelector('.drop-text').textContent = file.name;
  dropZone.querySelector('.drop-hint').textContent = (file.size/1024).toFixed(1)+' KB · 点击重新选择';
  uploadBtn.disabled = false;
}

function clearResult(){
  result.innerHTML = '';
  dropZone.classList.remove('has-file');
  dropZone.querySelector('.drop-icon').textContent = '⬆';
  dropZone.querySelector('.drop-text').textContent = '点击或拖拽文件上传';
  dropZone.querySelector('.drop-hint').textContent = '支持 PDF / JPG / PNG / TXT(扫描件自动 OCR)· 单文件 ≤ 10MB';
  fileInput.value = '';
  currentFile = null;
  uploadBtn.disabled = true;
}

document.getElementById('upload-form').addEventListener('submit', async e=>{
  e.preventDefault();
  if(!currentFile) return;
  uploadBtn.disabled = true;
  uploadBtn.innerHTML = '<span class="loading"></span>解析中...';
  result.innerHTML = '<div style="padding:24px;text-align:center;color:var(--muted-foreground);font-size:13px"><span class="loading"></span>OCR 提取 + AI 解析中(约 10-30 秒)...</div>';
  try{
    const fd = new FormData();
    fd.append('file', currentFile);
    const r = await fetch('/api/upload', {method:'POST', body:fd});
    const j = await r.json();
    if(j.ok){
      const d = j.result;
      const conf = d.confidence||0;
      const confClass = conf>=0.8?'conf-high':conf>=0.5?'conf-mid':'conf-low';
      const confColor = conf>=0.8?'var(--success)':conf>=0.5?'var(--warning)':'var(--danger)';
      let html = '<div class="result-card">';
      html += '<div class="result-head"><h3>解析完成</h3><span class="lvl-tag">'+escapeHtml(d.level1||'未归类')+' / '+escapeHtml(d.level2||'未归类')+'</span></div>';
      html += '<div class="field-grid">';
      html += '<div class="field"><span class="field-label">发票号码</span><span class="field-value">'+escapeHtml(d.invoice_no||'—')+'</span></div>';
      html += '<div class="field"><span class="field-label">开票日期</span><span class="field-value">'+escapeHtml(d.invoice_date||'—')+'</span></div>';
      html += '<div class="field"><span class="field-label">销售方</span><span class="field-value">'+escapeHtml(d.vendor||'—')+'</span></div>';
      html += '<div class="field"><span class="field-label">金额</span><span class="field-value amount">¥'+(d.amount||0).toLocaleString('zh-CN')+'</span></div>';
      html += '</div>';
      if(d.items && d.items.length){
        html += '<div class="field" style="margin-top:14px"><span class="field-label">商品明细</span><div class="items-list">'+d.items.map(it=>'<span class="item-tag">'+escapeHtml(typeof it==='string'?it:(it.name||it.description||JSON.stringify(it)))+'</span>').join('')+'</div></div>';
      }
      html += '<div class="field" style="margin-top:14px"><span class="field-label">AI 判断</span><span class="field-value">'+escapeHtml(d.reason||'—')+'</span>';
      html += '<div class="confidence-wrap"><div class="confidence-bar"><div class="confidence-fill" style="width:'+(conf*100)+'%;background:'+confColor+'"></div></div>';
      html += '<span class="confidence-label '+confClass+'">置信度 '+(conf*100).toFixed(0)+'%</span></div></div>';
      if(j.ocr_text_preview){
        html += '<details style="margin-top:14px"><summary>查看 OCR 原文</summary><div class="ocr-preview">'+escapeHtml(j.ocr_text_preview)+'</div></details>';
      }
      html += '<a class="next-link" href="/report">查看管报预览 →</a>';
      html += '</div>';
      result.innerHTML = html;
    } else {
      result.innerHTML = '<div class="error-box">解析失败:'+escapeHtml(j.error||JSON.stringify(j))+'</div>';
    }
  }catch(e){
    result.innerHTML = '<div class="error-box">网络错误:'+escapeHtml(e.message)+'</div>';
  }
  uploadBtn.disabled = false;
  uploadBtn.textContent = '开始解析';
});
</script>
</body>
</html>"""




# === 路由 ===
@app.route("/login", methods=["GET", "POST"])
def login_page():
    if request.method == "GET":
        return LOGIN_HTML.replace("{% if error %}", "").replace("{% endif %}", "").replace("{{ error }}", "")
    name = request.form.get("name", "").strip()
    pin = request.form.get("pin", "").strip()
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT id, name, role FROM users WHERE name=? AND pin=?", (name, pin))
        row = c.fetchone()
    finally:
        conn.close()
    if not row:
        return LOGIN_HTML.replace("{% if error %}", "").replace("{% endif %}", "").replace("{{ error }}", "姓名或 PIN 码错误")
    flask_session["user_id"] = row[0]
    return redirect("/")


@app.route("/logout")
def logout_page():
    flask_session.clear()
    return redirect("/login")


@app.route("/")
def index():
    u = _current_user()
    if not u:
        return redirect("/login")
    return INDEX_HTML

@app.route("/upload")
def upload_page():
    u = _current_user()
    if not u:
        return redirect("/login")
    return UPLOAD_HTML

@app.route("/health")
def health():
    return jsonify({
        "status": "ok",
        "service": "finance-agent",
        "day": "D4",
        "port": 5002,
        "llm_model": LLM_MODEL,
    })

# === API ===
@app.route("/api/test-llm")
def test_llm():
    """测试 CherryIN 连通性"""
    result = test_connection()
    return jsonify(result)

@app.route("/api/normalize", methods=["POST"])
def normalize():
    """字段归一化: 输入流水,输出归一化科目"""
    data = request.get_json(silent=True) or {}
    amount = data.get("amount", 0)
    try:
        amount = float(amount)
    except (ValueError, TypeError):
        amount = 0
    summary = data.get("summary", "")
    source = data.get("source", "未知")

    if not summary:
        return jsonify({"ok": False, "error": "summary is required"})

    prompt = NORMALIZE_PROMPT.format(
        rules=ACCOUNTING_RULES,
        amount=amount,
        summary=summary,
        source=source,
    )

    result = chat_json([
        {"role": "system", "content": "你是财务归类助手,严格按口径规则归一化流水科目。只返回 JSON。"},
        {"role": "user", "content": prompt},
    ], temperature=0.1)

    if isinstance(result, dict) and result.get("_error"):
        return jsonify({"ok": False, "error": result["_error"], "stage": "normalize"})
    if not isinstance(result, dict) or "level1" not in result:
        return jsonify({"ok": False, "error": "LLM 返回格式异常", "stage": "normalize", "raw": str(result)[:200]})
    return jsonify({"ok": True, "result": result, "input": {"amount": amount, "summary": summary, "source": source}})

@app.route("/api/normalize/batch", methods=["POST"])
def normalize_batch():
    """批量归一化"""
    data = request.get_json(silent=True) or {}
    transactions = data.get("transactions", [])
    if not isinstance(transactions, list):
        return jsonify({"ok": False, "error": "transactions must be a list"})
    results = []
    for tx in transactions:
        if not isinstance(tx, dict):
            results.append({"input": tx, "ok": False, "error": "invalid transaction"})
            continue
        amount = tx.get("amount", 0)
        try:
            amount = float(amount)
        except (ValueError, TypeError):
            amount = 0
        summary = tx.get("summary", "")
        source = tx.get("source", "未知")
        if not summary:
            results.append({"ok": False, "error": "no summary"})
            continue
        prompt = NORMALIZE_PROMPT.format(
            rules=ACCOUNTING_RULES, amount=amount, summary=summary, source=source,
        )
        r = chat_json([
            {"role": "system", "content": "你是财务归类助手。只返回 JSON。"},
            {"role": "user", "content": prompt},
        ], temperature=0.1)
        if isinstance(r, dict) and r.get("_error"):
            results.append({"input": tx, "ok": False, "error": r["_error"]})
        elif not isinstance(r, dict) or "level1" not in r:
            results.append({"input": tx, "ok": False, "error": "LLM 返回格式异常"})
        else:
            results.append({"input": tx, "ok": True, "result": r})
    return jsonify({"ok": True, "count": len(results), "success_count": len([r for r in results if r.get("ok")]), "results": results})

# === 数据上传 ===
def _ocr_pdf(pdf_bytes):
    """对扫描 PDF 做 OCR,返回 (text, error)。复用法务 Agent 的 OCR 实现。"""
    try:
        from pdf2image import convert_from_bytes
        import pytesseract
    except ImportError as e:
        return None, f"OCR 依赖未安装: {e}"
    try:
        images = convert_from_bytes(pdf_bytes, dpi=_OCR_DPI, first_page=1, last_page=_OCR_MAX_PAGES)
    except Exception as e:
        return None, f"PDF 转图片失败: {e}"
    if not images:
        return None, "PDF 无可渲染页面"
    pages_text = []
    for img in images:
        try:
            pages_text.append(pytesseract.image_to_string(img, lang="chi_sim+eng").strip())
        except Exception:
            pages_text.append("")
        img.close()
    text = "\n\n".join(t for t in pages_text if t)
    return text, None


def extract_text_from_upload(f, filename):
    """从上传文件提取文本。成功返回 str,失败返回 dict。
    支持: PDF(文本型/扫描型), 图片(png/jpg,直接 OCR), TXT"""
    raw = f.read()
    if not raw:
        return {"ok": False, "error": "文件为空"}

    fl = filename.lower()

    # --- 图片直接 OCR ---
    if fl.endswith((".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp")):
        try:
            import pytesseract
            from PIL import Image
            import io as _io
            img = Image.open(_io.BytesIO(raw))
            text = pytesseract.image_to_string(img, lang="chi_sim+eng").strip()
            if not text:
                return {"ok": False, "error": "图片 OCR 未提取到文本(图片可能模糊或无文字)"}
            return text
        except ImportError as e:
            return {"ok": False, "error": f"OCR 依赖未安装: {e}"}
        except Exception as e:
            return {"ok": False, "error": f"图片 OCR 失败: {e}"}

    # --- PDF: 先 pypdf,文本短则降级 OCR ---
    if fl.endswith(".pdf"):
        text = ""
        try:
            from pypdf import PdfReader
            import io as _io
            reader = PdfReader(_io.BytesIO(raw))
            text = "\n".join(page.extract_text() or "" for page in reader.pages)
        except ImportError:
            return {"ok": False, "error": "pypdf not installed"}
        except Exception as e:
            return {"ok": False, "error": f"PDF 解析失败: {e}"}

        if len(text.strip()) < _OCR_MIN_CHARS:
            ocr_text, ocr_err = _ocr_pdf(raw)
            if ocr_err:
                return {"ok": False, "error": f"扫描件需 OCR,但 OCR 失败: {ocr_err}"}
            if not ocr_text.strip():
                return {"ok": False, "error": "OCR 未提取到文本(图片可能模糊或为纯图形)"}
            return ocr_text
        return text

    # --- TXT ---
    if fl.endswith(".txt"):
        text = raw.decode("utf-8", errors="ignore")
        if not text.strip():
            return {"ok": False, "error": "文件为空"}
        return text

    return {"ok": False, "error": "不支持的文件格式(请上传 PDF / 图片 / TXT)"}


def parse_excel(file_storage):
    """解析 Excel/CSV,返回 (rows, error) 元组。成功时 error=None,失败时 rows=None"""
    filename = file_storage.filename
    if filename.endswith(".csv"):
        import csv
        stream = io.TextIOWrapper(file_storage.stream, encoding="utf-8-sig")
        reader = csv.DictReader(stream)
        rows = []
        for row in reader:
            # 自动找金额列和摘要列
            amount = 0
            summary = ""
            for k, v in row.items():
                if k and v:
                    kl = k.lower().strip()
                    if any(x in kl for x in ["金额", "amount", "amt", "总额", "费用"]):
                        try:
                            amount = float(str(v).replace(",", "").replace("¥", "").strip())
                        except (ValueError, TypeError):
                            pass
                    elif any(x in kl for x in ["摘要", "summary", "说明", "备注", "事由", "用途"]):
                        summary = str(v).strip()
            if summary or amount:
                rows.append({"summary": summary, "amount": amount, "source": "上传"})
        return rows, None
    else:
        try:
            import openpyxl
        except ImportError:
            return None, "openpyxl not installed"
        try:
            wb = openpyxl.load_workbook(file_storage.stream, data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                amount = 0
                summary = ""
                for k, v in row_dict.items():
                    if k and v is not None:
                        kl = k.lower().strip()
                        if any(x in kl for x in ["金额", "amount", "amt", "总额", "费用"]):
                            try:
                                amount = float(str(v).replace(",", "").replace("¥", "").strip())
                            except (ValueError, TypeError):
                                pass
                        elif any(x in kl for x in ["摘要", "summary", "说明", "备注", "事由", "用途"]):
                            summary = str(v).strip()
                if summary or amount:
                    rows.append({"summary": summary, "amount": amount, "source": "上传"})
            return rows, None
        except Exception as e:
            return None, f"Excel 解析失败: {e}"

@app.route("/api/upload", methods=["POST"])
def upload():
    """上传发票 PDF/图片,OCR 提取 → AI 解析+归一化 → 入库"""
    u = _current_user()
    if not u:
        return jsonify({"ok": False, "error": "未登录"}), 401
    f = request.files.get("file")
    if not f:
        return jsonify({"ok": False, "error": "file is required"})
    filename = f.filename or "upload.bin"

    # 1. 提取文本
    extracted = extract_text_from_upload(f, filename)
    if isinstance(extracted, dict) and not extracted.get("ok", True):
        return jsonify(extracted)
    invoice_text = extracted
    if len(invoice_text.strip()) < 5:
        return jsonify({"ok": False, "error": "提取到的文本过短,无法解析"})

    # 2. AI 解析 + 归一化(一步)
    prompt = INVOICE_PARSE_PROMPT.format(rules=ACCOUNTING_RULES, invoice_text=invoice_text[:6000])
    r = chat_json([
        {"role": "system", "content": "你是财务发票解析助手。从 OCR 文本提取字段并归一化科目。只返回 JSON。"},
        {"role": "user", "content": prompt},
    ], temperature=0.1)

    if not isinstance(r, dict) or r.get("_error") or "level1" not in r:
        return jsonify({"ok": False, "error": r.get("_error", "LLM 返回异常") if isinstance(r, dict) else "LLM 返回非 dict",
                         "stage": "parse", "ocr_text_preview": invoice_text[:500]})

    # 3. 解析金额(容错)
    try:
        amount = float(str(r.get("amount", 0)).replace(",", "").replace("¥", "").replace("元", "").strip() or 0)
    except (ValueError, TypeError):
        amount = 0

    level1 = r.get("level1", "未归类")
    level2 = r.get("level2", "未归类")
    confidence = max(0, min(1.0, float(r.get("confidence", 0))))
    reason = r.get("reason", "")
    vendor = r.get("vendor", "")
    invoice_no = r.get("invoice_no", "")
    invoice_date = r.get("invoice_date", "")
    items = r.get("items", [])
    summary = " / ".join(items) if items else (vendor or "发票")

    # 4. 入库
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute(
            "INSERT INTO transactions (source, amount, summary, level1, level2, confidence, reason, vendor, invoice_no, invoice_text, user_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
            ("发票", amount, summary, level1, level2, confidence, reason, vendor, invoice_no, invoice_text[:5000], flask_session.get("user_id")),
        )
        conn.commit()
        tx_id = c.lastrowid
    finally:
        conn.close()

    return jsonify({
        "ok": True,
        "transaction_id": tx_id,
        "result": {
            "invoice_no": invoice_no,
            "invoice_date": invoice_date,
            "vendor": vendor,
            "amount": amount,
            "items": items,
            "level1": level1,
            "level2": level2,
            "confidence": confidence,
            "reason": reason,
        },
        "ocr_text_preview": invoice_text[:500],
    })


@app.route("/api/report/preview")
def report_preview():
    """管报预览: 按一级科目汇总。财务可看全部,普通员工只看自己。"""
    u = _current_user()
    if not u:
        return jsonify({"ok": False, "error": "未登录"}), 401
    # 财务角色默认 all(看全部);普通员工默认且强制 mine(只看自己)
    if u["role"] == "financial":
        scope = request.args.get("scope", "all")
    else:
        scope = "mine"
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        if scope == "all":
            c.execute("SELECT level1, level2, COUNT(*) as cnt, SUM(amount) as total FROM transactions GROUP BY level1, level2 ORDER BY level1, level2")
            rows = c.fetchall()
            c.execute("SELECT COUNT(*) as total_count, SUM(amount) as total_amount FROM transactions")
            overall = c.fetchone()
        else:
            c.execute("SELECT level1, level2, COUNT(*) as cnt, SUM(amount) as total FROM transactions WHERE user_id=? GROUP BY level1, level2 ORDER BY level1, level2", (u["id"],))
            rows = c.fetchall()
            c.execute("SELECT COUNT(*) as total_count, SUM(amount) as total_amount FROM transactions WHERE user_id=?", (u["id"],))
            overall = c.fetchone()
    finally:
        conn.close()
    summary = {}
    for r in rows:
        level1, level2, cnt, total = r
        if level1 not in summary:
            summary[level1] = {"total": 0, "count": 0, "items": []}
        summary[level1]["total"] += total or 0
        summary[level1]["count"] += cnt
        summary[level1]["items"].append({"level2": level2, "count": cnt, "total": round(total or 0, 2)})
    # 取最近 10 条流水明细(按 scope 过滤)
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        if scope == "all":
            c.execute("""SELECT t.id, t.source, t.amount, t.summary, t.level1, t.level2,
                                t.vendor, t.invoice_no, t.confidence,
                                datetime(t.created_at,'+8 hours') as created_at,
                                u.name as user_name
                         FROM transactions t LEFT JOIN users u ON t.user_id=u.id
                         ORDER BY t.id DESC LIMIT 10""")
        else:
            c.execute("""SELECT t.id, t.source, t.amount, t.summary, t.level1, t.level2,
                                t.vendor, t.invoice_no, t.confidence,
                                datetime(t.created_at,'+8 hours') as created_at,
                                u.name as user_name
                         FROM transactions t LEFT JOIN users u ON t.user_id=u.id
                         WHERE t.user_id=? ORDER BY t.id DESC LIMIT 10""", (u["id"],))
        recent_rows = c.fetchall()
    finally:
        conn.close()
    recent = []
    for r in recent_rows:
        recent.append({
            "id": r[0], "source": r[1], "amount": round(r[2] or 0, 2),
            "summary": r[3] or "", "level1": r[4] or "未归类", "level2": r[5] or "",
            "vendor": r[6] or "", "invoice_no": r[7] or "",
            "confidence": r[8] or 0, "created_at": r[9] or "",
            "user_name": r[10] or "未绑定",
        })
    return jsonify({
        "ok": True,
        "total_count": overall[0],
        "total_amount": round(overall[1] or 0, 2),
        "summary": {k: {**v, "total": round(v["total"], 2)} for k, v in summary.items()},
        "recent": recent,
        "scope": scope,
    })

# === D4: 管报页面 + AI 简评 + 飞书输出 ===
# 简评缓存(进程级,报表 hash → commentary)
_COMMENTARY_CACHE = {}

def _get_report_data():
    """取管报数据,返回 (rows, overall) 或 (None, None)"""
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT level1, level2, COUNT(*) as cnt, SUM(amount) as total FROM transactions GROUP BY level1, level2 ORDER BY level1, level2")
        rows = c.fetchall()
        c.execute("SELECT COUNT(*) as total_count, SUM(amount) as total_amount FROM transactions")
        overall = c.fetchone()
    finally:
        conn.close()
    if not rows:
        return None, None
    return rows, overall

def _build_report_text(rows, overall):
    """构建管报文本(供 AI 简评用)"""
    lines = [f"总笔数: {overall[0]}, 总金额: ¥{overall[1] or 0:,.2f}"]
    cur_l1 = None
    for level1, level2, cnt, total in rows:
        if level1 != cur_l1:
            lines.append(f"\n【{level1}】")
            cur_l1 = level1
        lines.append(f"  {level2}: {cnt}笔, ¥{total or 0:,.2f}")
    return "\n".join(lines)

def _generate_commentary(rows, overall):
    """生成 AI 简评(带缓存,避免重复调 LLM)"""
    report_data = _build_report_text(rows, overall)
    cache_key = hash(report_data)
    if cache_key in _COMMENTARY_CACHE:
        return _COMMENTARY_CACHE[cache_key], None
    prompt = REPORT_COMMENTARY_PROMPT.format(report_data=report_data)
    result = chat([
        {"role": "system", "content": "你是财务分析师,正在给同事做管报简评。直接输出 3-5 行简评,每行一个要点。"},
        {"role": "user", "content": prompt},
    ], temperature=0.3)
    if result.get("error"):
        return None, result["error"]
    commentary = result.get("content", "").strip()
    if commentary:
        _COMMENTARY_CACHE[cache_key] = commentary
    return commentary, None

@app.route("/report")
def report_page():
    """管报预览页面"""
    u = _current_user()
    if not u:
        return redirect("/login")
    return REPORT_HTML

@app.route("/api/report/commentary", methods=["POST"])
def report_commentary():
    """AI 简评: 基于管报数据生成 3-5 条点评"""
    rows, overall = _get_report_data()
    if rows is None:
        return jsonify({"ok": False, "error": "暂无数据,请先上传"})
    commentary, err = _generate_commentary(rows, overall)
    if err:
        return jsonify({"ok": False, "error": f"AI 简评生成失败: {err}"})
    return jsonify({"ok": True, "commentary": commentary})

@app.route("/api/feishu/chats")
def feishu_chats():
    """列出 Bot 所在的飞书群聊"""
    return jsonify(feishu_list_chats())

@app.route("/api/report/feishu", methods=["POST"])
def report_feishu():
    """把管报 + AI 简评发到飞书群"""
    data = request.get_json(silent=True) or {}
    chat_id = data.get("chat_id", "")
    if not chat_id:
        return jsonify({"ok": False, "error": "chat_id is required"})

    rows, overall = _get_report_data()
    if rows is None:
        return jsonify({"ok": False, "error": "暂无数据"})

    # 构建管报富文本
    paragraphs = [[
        {"tag": "text", "text": f"总笔数 {overall[0]}, 总金额 ¥{overall[1] or 0:,.2f}\n"},
    ]]
    cur_l1 = None
    cur_items = []
    for level1, level2, cnt, total in rows:
        if level1 != cur_l1:
            if cur_items:
                paragraphs.append(cur_items)
            cur_l1 = level1
            cur_items = [{"tag": "text", "text": f"\n【{level1}】\n"}]
        cur_items.append({"tag": "text", "text": f"  {level2}: {cnt}笔 ¥{total or 0:,.2f}\n"})
    if cur_items:
        paragraphs.append(cur_items)

    # 1. 发送管报
    r1 = feishu_send_post(chat_id, "📊 管报预览", paragraphs)
    if not r1.get("ok"):
        return jsonify({"ok": False, "error": f"管报发送失败: {r1.get('error', 'unknown')}", "report_sent": False})

    # 2. 生成 AI 简评(带缓存,不重复调 LLM)
    commentary, err = _generate_commentary(rows, overall)
    if err or not commentary:
        return jsonify({"ok": True, "report_sent": True, "commentary_sent": False, "commentary_error": err or "empty commentary"})

    # 3. 发送简评
    comm_paragraphs = [[{"tag": "text", "text": commentary}]]
    r2 = feishu_send_post(chat_id, "🤖 AI 简评", comm_paragraphs)
    if not r2.get("ok"):
        return jsonify({"ok": True, "report_sent": True, "commentary_sent": False, "commentary": commentary, "commentary_error": r2.get("error", "unknown")})
    return jsonify({"ok": True, "report_sent": True, "commentary_sent": True, "commentary": commentary})

# === D5: 绩效试算 ===
PERFORMANCE_PROMPT = """你是绩效管理专家。根据以下数据给出绩效简评(2-3 条):

{data}

要求:
1. 每条一行,直接给结论
2. 关注达成率、超额/未达标、系数合理性
3. 不要客套话
"""

def _get_performance_rules():
    """取所有绩效规则,返回 list of dict"""
    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        c.execute("SELECT department, position, coefficient, target_amount, bonus_base FROM performance_rules ORDER BY department, position")
        rows = c.fetchall()
    finally:
        conn.close()
    return [{"department": r[0], "position": r[1], "coefficient": r[2],
             "target_amount": r[3], "bonus_base": r[4]} for r in rows]


def _calculate_performance(department=None):
    """计算绩效。返回 dict: {departments: [...], summary: {...}}

    逻辑:
      - 每个部门的流水 = transactions 表里 source 匹配部门名的金额合计
        (source 字段目前是"报销/对公支付/工资",不直接区分部门,
         所以用 LLM 归一化后的 level2 项关联部门关键词)
      - 达成率 = 部门实际流水 / target_amount
      - 绩效奖金 = bonus_base × coefficient × (达成率 clamp 到 [0, 1.5])
    """
    rules = _get_performance_rules()
    if not rules:
        return None

    conn = sqlite3.connect(DB_PATH)
    try:
        c = conn.cursor()
        # 按一级科目汇总,用于关联部门
        c.execute("SELECT level1, SUM(amount) FROM transactions GROUP BY level1")
        l1_totals = {r[0]: r[1] or 0 for r in c.fetchall()}
        c.execute("SELECT COUNT(*) as cnt, SUM(amount) as total FROM transactions")
        overall = c.fetchone()
    finally:
        conn.close()

    total_amount = overall[1] or 0
    total_count = overall[0] or 0

    # 部门 → 科目映射(简化模型:哪类费用占比高就归哪个部门)
    DEPT_SUBJECT_MAP = {
        "研发部": ["研发费"],
        "销售部": ["销售费"],
        "管理部": ["管理费"],
        "交付部": ["营业成本"],
    }

    results = []
    for rule in rules:
        dept = rule["department"]
        # 部门实际流水 = 映射科目金额合计(简化模型)
        subjects = DEPT_SUBJECT_MAP.get(dept, [])
        actual = sum(l1_totals.get(s, 0) for s in subjects)

        target = rule["target_amount"]
        achievement_rate = (actual / target) if target > 0 else 0
        # 达成率 clamp [0, 1.5]
        clamped_rate = max(0, min(1.5, achievement_rate))
        bonus = rule["bonus_base"] * rule["coefficient"] * clamped_rate

        status = "达标"
        if achievement_rate < 0.6:
            status = "未达标"
        elif achievement_rate >= 1.0:
            status = "超额"

        results.append({
            "department": dept,
            "position": rule["position"],
            "coefficient": rule["coefficient"],
            "target_amount": target,
            "actual_amount": actual,
            "achievement_rate": round(achievement_rate * 100, 1),
            "bonus_base": rule["bonus_base"],
            "bonus": round(bonus, 2),
            "status": status,
        })

    if department:
        results = [r for r in results if r["department"] == department]

    total_bonus = sum(r["bonus"] for r in results)

    return {
        "rules_count": len(rules),
        "total_count": total_count,
        "total_amount": total_amount,
        "total_bonus": round(total_bonus, 2),
        "departments": results,
        "level1_totals": l1_totals,
    }


@app.route("/performance")
def performance_page():
    """绩效试算页面"""
    return PERFORMANCE_HTML


@app.route("/api/performance/calculate", methods=["GET"])
def performance_calculate():
    """计算绩效试算结果"""
    department = request.args.get("department", "").strip()
    data = _calculate_performance(department or None)
    if data is None:
        return jsonify({"ok": False, "error": "未配置绩效规则"})
    return jsonify({"ok": True, **data})


@app.route("/api/performance/feishu", methods=["POST"])
def performance_send_feishu():
    """把绩效报告推送到飞书群"""
    data = request.get_json(silent=True) or {}
    chat_id = data.get("chat_id", "").strip()
    if not chat_id:
        return jsonify({"ok": False, "error": "缺少 chat_id"})

    dept = data.get("department", "").strip() or None
    result = _calculate_performance(dept)
    if result is None:
        return jsonify({"ok": False, "error": "未配置绩效规则"})

    # 构建飞书富文本
    title = "📊 绩效试算报告"
    paragraphs = [
        [{"tag": "text", "text": f"总流水: ¥{result['total_amount']:,.2f} ({result['total_count']} 笔)\n绩效奖金合计: ¥{result['total_bonus']:,.2f}\n"}],
    ]

    # 按部门分组
    for item in result["departments"]:
        status_emoji = {"达标": "✅", "超额": "🚀", "未达标": "⚠️"}.get(item["status"], "")
        line = f"{status_emoji} {item['department']} · {item['position']}\n  目标: ¥{item['target_amount']:,.0f} | 实际: ¥{item['actual_amount']:,.0f} | 达成率: {item['achievement_rate']}%\n  系数: {item['coefficient']} | 奖金基数: ¥{item['bonus_base']:,.0f} → 应发: ¥{item['bonus']:,.2f}"
        paragraphs.append([{"tag": "text", "text": line}])

    r = feishu_send_post(chat_id, title, paragraphs)
    if not r.get("ok"):
        return jsonify({"ok": False, "error": r.get("error", "发送失败")})
    return jsonify({"ok": True, "sent": True, "total_bonus": result["total_bonus"]})


@app.route("/api/performance/rules", methods=["GET"])
def performance_rules_api():
    """列出绩效规则"""
    return jsonify({"ok": True, "rules": _get_performance_rules()})


# === 飞书 webhook ===
# 已处理消息 ID 去重(飞书会重试)
_PROCESSED_MSG_IDS = set()
_MAX_MSG_CACHE = 200


def _handle_feishu_file_message(msg_id, msg_type, content, chat_id, sender_open_id):
    """处理飞书图片/文件消息:下载 → OCR → AI 解析 → 入库 → 回复"""
    print(f"[FILE_HANDLER] start msg_id={msg_id} type={msg_type} content_keys={list(content.keys())}", flush=True)
    try:
        # 提取 file_key / image_key
        if msg_type == "image":
            file_key = content.get("image_key", "")
            file_type = "image"
            filename = f"feishu_image_{file_key[:16]}.png"
        elif msg_type == "file":
            file_key = content.get("file_key", "")
            file_type = "file"
            filename = content.get("file_name", f"feishu_file_{file_key[:16]}")
        else:
            return

        if not file_key:
            feishu_send_text(chat_id, "❌ 无法获取文件 key")
            return

        feishu_send_text(chat_id, "📥 正在下载文件...")

        # 1. 下载文件
        dl = feishu_download_file(msg_id, file_key, file_type=file_type)
        if not dl.get("ok"):
            feishu_send_text(chat_id, f"❌ 文件下载失败: {dl.get('error', '未知错误')}")
            return

        file_bytes = dl["data"]
        feishu_send_text(chat_id, f"📄 文件已下载({len(file_bytes)} 字节),OCR 提取中...")

        # 2. 提取文本(复用 extract_text_from_upload)
        import io as _io
        extracted = extract_text_from_upload(_io.BytesIO(file_bytes), filename)
        if isinstance(extracted, dict) and not extracted.get("ok", True):
            feishu_send_text(chat_id, f"❌ 文本提取失败: {extracted.get('error', '未知')}")
            return
        invoice_text = extracted
        if len(invoice_text.strip()) < 5:
            feishu_send_text(chat_id, "❌ 提取到的文本过短,无法解析(图片可能模糊或无文字)")
            return

        # 3. AI 解析 + 归一化
        feishu_send_text(chat_id, "🤖 AI 解析中(约 10-30 秒)...")
        prompt = INVOICE_PARSE_PROMPT.format(rules=ACCOUNTING_RULES, invoice_text=invoice_text[:6000])
        r = chat_json([
            {"role": "system", "content": "你是财务发票解析助手。从 OCR 文本提取字段并归一化科目。只返回 JSON。"},
            {"role": "user", "content": prompt},
        ], temperature=0.1)

        if not isinstance(r, dict) or r.get("_error") or "level1" not in r:
            feishu_send_text(chat_id, f"❌ AI 解析失败: {r.get('_error', '返回异常') if isinstance(r, dict) else '非 dict'}")
            return

        # 4. 解析金额
        try:
            amount = float(str(r.get("amount", 0)).replace(",", "").replace("¥", "").replace("元", "").strip() or 0)
        except (ValueError, TypeError):
            amount = 0

        level1 = r.get("level1", "未归类")
        level2 = r.get("level2", "未归类")
        confidence = max(0, min(1.0, float(r.get("confidence", 0))))
        reason = r.get("reason", "")
        vendor = r.get("vendor", "")
        invoice_no = r.get("invoice_no", "")
        items = r.get("items", [])
        summary = " / ".join(items) if items else (vendor or "发票")

        # 5. 映射到 user(飞书 open_id → users 表)
        user = _get_or_create_user_by_open_id(sender_open_id)

        # 6. 入库
        conn = sqlite3.connect(DB_PATH)
        try:
            c = conn.cursor()
            c.execute(
                "INSERT INTO transactions (source, amount, summary, level1, level2, confidence, reason, vendor, invoice_no, invoice_text, user_id) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                ("发票", amount, summary, level1, level2, confidence, reason, vendor, invoice_no, invoice_text[:5000], user["id"] if user else None),
            )
            conn.commit()
        finally:
            conn.close()

        # 7. 回复结果
        conf_emoji = "✅" if confidence >= 0.8 else "⚠️" if confidence >= 0.5 else "❓"
        lines = [
            f"✅ 发票已解析入库",
            "",
            f"发票号: {invoice_no or '—'}",
            f"销售方: {vendor or '—'}",
            f"金额: ¥{amount:,.2f}",
            f"科目: {level1} / {level2}",
            f"{conf_emoji} 置信度: {confidence*100:.0f}%",
            f"理由: {reason}",
        ]
        if user:
            lines.append(f"归属: {user['name']}")
        lines.append("")
        lines.append('发送"管报"查看汇总')
        feishu_send_text(chat_id, "\n".join(lines))

    except Exception as e:
        try:
            feishu_send_text(chat_id, f"❌ 处理文件时出错: {e}")
        except Exception:
            pass


def _handle_feishu_message(text, chat_id):
    """处理飞书消息指令,异步调用(不阻塞 webhook 响应)"""
    print(f"[MSG_HANDLER] text={text}", flush=True)
    text = (text or "").strip()
    try:
        if "帮助" in text or text.lower() in ("help", "?", "？"):
            feishu_send_text(chat_id,
                "🤖 财务试算助手 · 指令列表\n"
                "──────────────\n"
                "管报  — 查看最新管报汇总\n"
                "绩效  — 查看绩效试算结果\n"
                "简评  — AI 生成财务简评\n"
                "帮助  — 显示本指令列表\n"
                "──────────────\n"
                "直接发送关键词即可,无需@")

        elif "管报" in text:
            rows, overall = _get_report_data()
            if rows is None:
                feishu_send_text(chat_id, "📊 暂无流水数据,请先上传 Excel/CSV。")
                return
            report = _build_report_text(rows, overall)
            feishu_send_text(chat_id, "📊 最新管报汇总\n" + report)

        elif "绩效" in text:
            data = _calculate_performance()
            if data is None:
                feishu_send_text(chat_id, "🎯 暂无绩效规则配置。")
                return
            lines = [
                f"🎯 绩效试算结果",
                f"总流水: ¥{data['total_amount']:,.2f} ({data['total_count']} 笔)",
                f"绩效奖金合计: ¥{data['total_bonus']:,.2f}",
                "",
            ]
            for item in data["departments"]:
                emoji = {"达标": "✅", "超额": "🚀", "未达标": "⚠️"}.get(item["status"], "")
                lines.append(f"{emoji} {item['department']}·{item['position']}: 达成率 {item['achievement_rate']}% → ¥{item['bonus']:,.0f}")
            feishu_send_text(chat_id, "\n".join(lines))

        elif "简评" in text:
            rows, overall = _get_report_data()
            if rows is None:
                feishu_send_text(chat_id, "🤖 暂无流水数据,无法生成简评。")
                return
            feishu_send_text(chat_id, "🤖 AI 正在分析管报,请稍候...")
            commentary, err = _generate_commentary(rows, overall)
            if err:
                feishu_send_text(chat_id, f"❌ 简评生成失败: {err}")
            else:
                feishu_send_text(chat_id, "🤖 AI 简评\n" + commentary)

        else:
            feishu_send_text(chat_id,
                f"收到: {text[:50]}\n发送\"帮助\"查看可用指令。")
    except Exception as e:
        try:
            feishu_send_text(chat_id, f"❌ 处理消息时出错: {e}")
        except Exception:
            pass


@app.route("/webhook", methods=["POST"])
def webhook():
    """飞书事件订阅回调

    支持飞书 v2 事件格式(header.event_type)和 v1 格式(event.type)。
    支持 Encrypt Key 加密模式。
    收到消息后立即返回 200,异步处理指令(避免飞书 3 秒超时)。
    """
    data = request.get_json(silent=True) or {}

    # 加密模式:飞书发 {"encrypt": "base64..."},需解密
    if "encrypt" in data and not data.get("challenge"):
        try:
            import base64, hashlib
            from Crypto.Cipher import AES
            encrypt_key = os.environ.get("LARK_ENCRYPT_KEY", "")
            if not encrypt_key:
                return jsonify({"error": "encrypt key not configured"}), 500
            key = hashlib.sha256(encrypt_key.encode("utf-8")).digest()
            enc = base64.b64decode(data["encrypt"])
            cipher = AES.new(key, AES.MODE_CBC, iv=enc[:16])
            decrypted = cipher.decrypt(enc[16:])
            pad = decrypted[-1]
            decrypted = decrypted[:-pad].decode("utf-8")
            data = json.loads(decrypted)
        except Exception as e:
            return jsonify({"error": f"decrypt failed: {e}"}), 500

    # challenge 验证(必须在 1 秒内返回)
    if "challenge" in data:
        return jsonify({"challenge": data["challenge"]})

    # 解析消息(v2 和 v1 兼容)
    header = data.get("header", {})
    event_type = header.get("event_type") or data.get("type", "")
    event = data.get("event", {})
    msg = event.get("message", {})

    if not msg:
        return jsonify({"ok": True})

    # 消息去重(飞书会重试)
    msg_id = msg.get("message_id", "")
    if msg_id and msg_id in _PROCESSED_MSG_IDS:
        return jsonify({"ok": True, "dedup": True})
    if msg_id:
        _PROCESSED_MSG_IDS.add(msg_id)
        if len(_PROCESSED_MSG_IDS) > _MAX_MSG_CACHE:
            _PROCESSED_MSG_IDS.pop()

    chat_id = msg.get("chat_id", "")
    content_str = msg.get("content", "{}")
    try:
        content = json.loads(content_str) if isinstance(content_str, str) else content_str
    except (json.JSONDecodeError, TypeError):
        content = {}
    text = content.get("text", "")

    # 获取消息类型和发送者
    msg_type = msg.get("message_type", "")
    sender = event.get("sender", {}).get("sender_id", {}).get("open_id", "")

    print(f"[WEBHOOK] msg_type={msg_type} chat_id={chat_id} sender={sender} text={text[:50] if text else ''}", flush=True)
    # 异步处理(不阻塞 webhook 响应)
    if chat_id:
        import threading
        if msg_type in ("image", "file"):
            # 图片/文件消息 → 发票上传流程
            t = threading.Thread(
                target=_handle_feishu_file_message,
                args=(msg_id, msg_type, content, chat_id, sender),
                daemon=True,
            )
            t.start()
        elif text:
            # 文本消息 → 指令处理
            t = threading.Thread(target=_handle_feishu_message, args=(text, chat_id), daemon=True)
            t.start()

    return jsonify({"ok": True})


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5002)), debug=os.environ.get("FLASK_DEBUG") == "1")
